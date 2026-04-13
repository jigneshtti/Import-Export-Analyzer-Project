import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font


def ensure_output_folder():
    os.makedirs("output", exist_ok=True)


def plot_vendor_chart(vendor_df):
    ensure_output_folder()

    plt.figure()
    plt.bar(vendor_df['vendor'], vendor_df['amount'])

    plt.title("Revenue by Vendor")
    plt.xlabel("Vendor")
    plt.ylabel("Revenue")
    plt.xticks(rotation=45)

    plt.tight_layout()
    plt.savefig("output/vendor_chart.png")
    plt.close()


def export_to_excel(summary, vendor_df):
    ensure_output_folder()

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Metric"
    ws["B1"] = "Value"

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    row = 2
    for key, value in summary.items():
        ws[f"A{row}"] = key.replace("_", " ").title()
        ws[f"B{row}"] = value
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    ws2 = wb.create_sheet(title="Vendor Summary")

    ws2["A1"] = "Vendor"
    ws2["B1"] = "Revenue"

    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)

    row = 2
    for _, data in vendor_df.iterrows():
        ws2[f"A{row}"] = data['vendor']
        ws2[f"B{row}"] = data['amount']
        row += 1

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    file_path = "output/report.xlsx"
    wb.save(file_path)

    return file_path
